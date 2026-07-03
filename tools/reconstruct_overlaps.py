#!/usr/bin/env python3
"""
reconstruct_overlaps.py
=======================
Reverse-engineers the "overlap input sheet" for BC Crown land tenures by
intersecting each tenure's Tantalis geometry against the survey-parcel fabric,
then derives WITHIN_Y_N from an area-coverage ratio and detects any unsurveyed
remainder. One reconstructed sheet per tenure is written to a single .xlsx whose
format matches multi_shape.xlsx, so the results drop straight into the LDS
generator's test harness.

It also performs a THRESHOLD CALIBRATION sweep: for each tenure it parses the
already-validated TENURE_LEGAL_DESCRIPTION to recover which parcels the human
author treated as whole vs partial, then finds the coverage threshold that best
reproduces those human splits. That threshold is the parameter the future GIS
automation should hardcode for WITHIN_Y_N.

INPUTS  : Query_2_result.csv  (INTRID_SID + validated LEGAL_DESC, from earlier pull)
OUTPUTS : reconstructed_overlaps.xlsx     (one sheet per tenure)
          reconstruction_report.csv       (per-tenure diagnostics)
          threshold_calibration.csv        (agreement % per candidate threshold)

  >>> EVERYTHING ENVIRONMENT-SPECIFIC IS IN THE CONFIG BLOCK BELOW. <<<
  Verification status of the CONFIG names: the survey-parcel fabric block was
  VERIFIED against the TA_SURVEY_PARCELS_SVW schema during discovery. The tenure
  and land-district blocks carry the standard catalogue names but were only
  exercised, not formally checked — `DESC` them before reusing in production
  code (the port keeps its own status list in src/lds/spatial/views.py).
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field

import oracledb
from openpyxl import Workbook

# =====================================================================
# CONFIG  — correct these against your catalog (DESC the views first)
# =====================================================================
CONFIG = {
    # Same env-var names as the production stage (src/lds/spatial/db.py).
    "dsn":      os.getenv('BCGW_DSN'),
    "user":     os.getenv('BCGW_USER'),
    "password": os.getenv('BCGW_PASSWORD'),

    # --- Tenure geometry source (Tantalis) ---
    "tenure_view":      "WHSE_TANTALIS.TA_CROWN_TENURES_SVW",
    "tenure_id_col":    "INTRID_SID",
    "tenure_geom_col":  "SHAPE",
    "tenure_area_col":  "TENURE_AREA_IN_HECTARES",

    # --- Parcel fabric source (survey parcels carry the Crown legal strings) ---
    # Column names VERIFIED against TA_SURVEY_PARCELS_SVW schema.
    "parcel_view":      "WHSE_TANTALIS.TA_SURVEY_PARCELS_SVW",
    "parcel_geom_col":  "SHAPE",
    "parcel_legal_col": "PARCEL_LEGAL_DESCRIPTION",   # the full legal text
    "parcel_id_col":    "PIN_SID",                    # unique survey-parcel id
    "parcel_type_col":  "PARCEL_TYPE",                # Primary / Subdivision / Right-of-Way
    "parcel_pid_col":   "LAND_TITLE_OFFICE_IDENTIFIER",  # non-null => titled (PID)

    # Filters (toggle to match the GIS workflow; confirm intent with Allan):
    "titled_only":      False,   # True => WHERE PID IS NOT NULL (titled parcels only)
    "exclude_row":      True,    # True => drop PARCEL_TYPE = 'Right-of-Way' (out of v1 scope)

    # --- Land districts (for the unsurveyed remainder's district) ---
    "ld_view":          "WHSE_TANTALIS.TA_LAND_DISTRICTS_SVW",
    "ld_name_col":      "LAND_DISTRICT_NAME",
    "ld_geom_col":      "SHAPE",

    # --- Geometry params ---
    "tolerance":        0.005,      # SDO tolerance in layer units (BC Albers metres)
    "sliver_sqm":       1.0,        # ignore parcel overlaps below this area (m^2)
    "remainder_ha_min": 0.01,       # remainder above this => unsurveyed component present

    # --- WITHIN_Y_N default threshold (aligned with the production
    #     WHOLE_PART_THRESHOLD in src/lds/constants.py; --threshold overrides) ---
    "within_threshold": 0.999,

    # --- Run control ---
    "max_workers":      8,          # parallel Oracle sessions (one per tenure slice)
}

# Working directory for inputs/outputs. Set LDS_DISCOVERY_DIR to the team
# discovery share (path recorded locally in docs/data_locations.md); defaults
# to the current directory. --input/--tag override individual paths.
_WORK_DIR = os.getenv("LDS_DISCOVERY_DIR", ".")
CONFIG.update({
    "input_csv":  os.path.join(_WORK_DIR, "calibration_candidates.csv"),
    "out_xlsx":   os.path.join(_WORK_DIR, "reconstructed_overlaps_calib2.xlsx"),
    "out_report": os.path.join(_WORK_DIR, "reconstruction_report_calib2.csv"),
    "out_calib":  os.path.join(_WORK_DIR, "threshold_calibration_calib2.csv"),
})

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("reconstruct")


# =====================================================================
# SQL — parameterized by a single tenure's INTRID_SID (:pid)
# =====================================================================
def parcel_overlap_sql(c: dict) -> str:
    """One row per survey parcel interacting with the tenure, plus the
    coverage ratio that drives WITHIN_Y_N."""
    # Optional fabric filters, confirmed against PARCEL_TYPE / PID columns.
    filt = []
    if c.get("exclude_row"):
        filt.append(f"p.{c['parcel_type_col']} <> 'Right-of-Way'")
    if c.get("titled_only"):
        filt.append(f"p.{c['parcel_pid_col']} IS NOT NULL")
    filt_sql = ("AND " + " AND ".join(filt)) if filt else ""
    return f"""
        WITH tenure AS (
            SELECT {c['tenure_geom_col']} AS geom,
                   {c['tenure_area_col']} AS tenure_area_ha
            FROM   {c['tenure_view']}
            WHERE  {c['tenure_id_col']} = :pid
        ),
        cand AS (
            SELECT p.{c['parcel_id_col']}    AS parcel_id,
                   p.{c['parcel_legal_col']} AS legal_desc,
                   p.{c['parcel_type_col']}  AS parcel_type,
                   p.{c['parcel_pid_col']}   AS pid,
                   p.{c['parcel_geom_col']}  AS pgeom,
                   t.geom                    AS tgeom,
                   t.tenure_area_ha          AS tenure_area_ha
            FROM   tenure t
            JOIN   {c['parcel_view']} p
                   ON SDO_ANYINTERACT(p.{c['parcel_geom_col']}, t.geom) = 'TRUE'
                   {filt_sql}
        ),
        meas AS (
            SELECT parcel_id, legal_desc, parcel_type, pid, tenure_area_ha,
                   SDO_GEOM.SDO_AREA(pgeom, :tol, 'unit=SQ_METER') AS parcel_sqm,
                   SDO_GEOM.SDO_AREA(
                       SDO_GEOM.SDO_INTERSECTION(pgeom, tgeom, :tol),
                       :tol, 'unit=SQ_METER') AS overlap_sqm
            FROM cand
        )
        SELECT legal_desc,
               parcel_type,
               pid,
               ROUND(parcel_sqm, 2)  AS parcel_sqm,
               ROUND(overlap_sqm, 2) AS overlap_sqm,
               ROUND(overlap_sqm / NULLIF(parcel_sqm, 0), 6) AS coverage_ratio,
               tenure_area_ha
        FROM   meas
        WHERE  overlap_sqm > :sliver
        ORDER BY coverage_ratio DESC
    """


def remainder_sql(c: dict) -> str:
    """Hectares of the tenure left over after subtracting all interacting
    parcels => presence of unsurveyed upland/foreshore."""
    filt = []
    if c.get("exclude_row"):
        filt.append(f"p.{c['parcel_type_col']} <> 'Right-of-Way'")
    if c.get("titled_only"):
        filt.append(f"p.{c['parcel_pid_col']} IS NOT NULL")
    filt_sql = ("AND " + " AND ".join(filt)) if filt else ""
    return f"""
        WITH tenure AS (
            SELECT {c['tenure_geom_col']} AS geom
            FROM   {c['tenure_view']}
            WHERE  {c['tenure_id_col']} = :pid
        ),
        punion AS (
            SELECT SDO_AGGR_UNION(SDOAGGRTYPE(p.{c['parcel_geom_col']}, :tol)) AS pgeom
            FROM   tenure t
            JOIN   {c['parcel_view']} p
                   ON SDO_ANYINTERACT(p.{c['parcel_geom_col']}, t.geom) = 'TRUE'
                   {filt_sql}
        )
        SELECT ROUND(
                 SDO_GEOM.SDO_AREA(
                   CASE WHEN pu.pgeom IS NULL THEN t.geom
                        ELSE SDO_GEOM.SDO_DIFFERENCE(t.geom, pu.pgeom, :tol) END,
                   :tol, 'unit=HECTARE'), 4) AS remainder_ha
        FROM   tenure t, punion pu
    """


def remainder_district_sql(c: dict) -> str:
    """Land district covering the largest share of the tenure — used as the
    unsurveyed remainder's district (approx; real rule may default to parcels')."""
    return f"""
        WITH tenure AS (
            SELECT {c['tenure_geom_col']} AS geom
            FROM   {c['tenure_view']}
            WHERE  {c['tenure_id_col']} = :pid
        )
        SELECT ld.{c['ld_name_col']} AS district,
               SDO_GEOM.SDO_AREA(
                   SDO_GEOM.SDO_INTERSECTION(ld.{c['ld_geom_col']}, t.geom, :tol),
                   :tol, 'unit=SQ_METER') AS overlap_sqm
        FROM   tenure t
        JOIN   {c['ld_view']} ld
               ON SDO_ANYINTERACT(ld.{c['ld_geom_col']}, t.geom) = 'TRUE'
        ORDER BY overlap_sqm DESC
        FETCH FIRST 1 ROW ONLY
    """


# =====================================================================
# Validated-description parser  (for threshold calibration)
# =====================================================================
# Recovers, per parcel NUMBER, whether the human author treated it as WHOLE
# or PARTIAL: {parcel_number: 'whole'|'partial'}. Calibration aligns each
# geometry parcel (by number) to the author's treatment.
#
# HARDENED against failure modes found in real Tantalis text:
#   - plan numbers ("PLAN 1192", "PLAN A164") must NOT be scraped as parcels
#   - township/range/meridian digits must NOT be scraped as parcels
#   - "those portions of" / "portion of" are PARTIAL markers (not just "part of")
#   - runtogether clauses with no delimiter still split on the markers
#   - REMAINDER / fractional-only descriptions can't be modeled by parcel number
#     -> return UNPARSEABLE so calibration SKIPS them rather than inventing numbers
#
# parse_parcel_treatment returns either a dict, or the sentinel UNPARSEABLE.
UNPARSEABLE = object()

# Partial-overlap markers (now incl. "portion(s)").
PART_MARKER_RE = re.compile(
    r'(those parts of|that part of|those portions of|that portion of|portions of|portion of)',
    re.IGNORECASE,
)

# A parcel reference: a container keyword + its number run. Crucially this only
# fires for true parcel containers, and a negative lookbehind blocks "PLAN <n>".
# Number run handles "525S, 527S and 2346S" style enumerations.
PARCEL_REF_RE = re.compile(
    r'(?<!plan )'                                   # not a plan number
    r'\b(?:district lots?|d\.?l\.?s?|lots?|sections?|sublots?)\s+'
    r'([0-9]+[a-z]?(?:\s*,\s*[0-9]+[a-z]?)*(?:\s*,?\s*(?:and|&)\s+[0-9]+[a-z]?)?)',
    re.IGNORECASE,
)

# Signals that the description is NOT modelable as numbered whole/partial parcels.
# REMAINDER arithmetic, fractional-only refs, and multi-clause legal prose.
UNMODELABLE_RE = re.compile(
    r'\bREMAINDER\b|\bREM\.|'                        # remainder-of-quarter arithmetic
    r'\bFirstly\b|\bSecondly\b|\bThirdly\b',         # multi-clause legal structure
    re.IGNORECASE,
)


def parse_parcel_treatment(desc: str):
    """Return {normalized_parcel_number: 'whole'|'partial'}, or UNPARSEABLE if the
    description can't be reduced to a numbered-parcel model. Calibration-only."""
    if not desc:
        return UNPARSEABLE
    text = re.sub(r'\s+', ' ', desc.strip())

    # Bail out cleanly on descriptions the numbered-parcel model can't represent,
    # rather than hallucinating numbers from REMAINDER/township/plan digits.
    if UNMODELABLE_RE.search(text):
        return UNPARSEABLE

    # Split into (span, treatment). Markers may abut prior text (no delimiter),
    # so we split on the marker itself; leading text before the first marker is
    # WHOLE, every marker-led clause is PARTIAL.
    spans = []
    markers = list(PART_MARKER_RE.finditer(text))
    if not markers:
        spans.append((text, 'whole'))
    else:
        if markers[0].start() > 0:
            spans.append((text[:markers[0].start()], 'whole'))
        for i, m in enumerate(markers):
            end = markers[i + 1].start() if i + 1 < len(markers) else len(text)
            spans.append((text[m.start():end], 'partial'))

    treatment = {}
    for span_text, treat in spans:
        for num in _parcels_in_span(span_text):
            # partial wins if a number somehow appears in both spans
            if num not in treatment or treat == 'partial':
                treatment[num] = treat

    # If after all that we found no real parcels, treat as unparseable so it's
    # skipped (e.g. a description that was entirely fractional/plan refs).
    return treatment if treatment else UNPARSEABLE


def _parcels_in_span(span_text: str) -> list:
    """Extract normalized parcel numbers from a span, guarding against plan
    numbers and township/range digits."""
    out = []
    for m in PARCEL_REF_RE.finditer(span_text):
        run = m.group(1)
        # Township/range/plan digits are guarded by the container keyword itself:
        # PARCEL_REF_RE only matches after lot/section/sublot/dl, never after
        # township/range, and a negative lookbehind blocks "PLAN <n>". No extra
        # look-ahead guard is applied.
        for tok in re.split(r',|\band\b|&', run):
            num = _norm_num(tok)
            if not num:
                continue
            out.append(num)
    return out


def _norm_num(tok: str) -> str:
    """Normalize a parcel-number token (' 526S ' -> '526S')."""
    t = tok.strip().upper()
    t = re.sub(r'[^0-9A-Z]', '', t)
    # Drop empties and tokens that are only letters (e.g. stray 'S' from a split).
    if not t or not re.search(r'[0-9]', t):
        return ''
    return t


def extract_parcel_numbers(legal_desc: str) -> set:
    """Parcel number(s) from a reconstructed parcel's PARCEL_LEGAL_DESCRIPTION,
    same plan-number / grid guards as the author-side parser."""
    if not legal_desc:
        return set()
    nums = set()
    for num in _parcels_in_span(re.sub(r'\s+', ' ', legal_desc)):
        nums.add(num)
    return nums


# =====================================================================
# Per-tenure worker
# =====================================================================
@dataclass
class TenureResult:
    intrid: str
    validated_desc: str
    parcels: list = field(default_factory=list)   # list of dicts
    tenure_area_ha: float | None = None
    remainder_ha: float | None = None
    remainder_district: str | None = None
    error: str | None = None


def process_tenure(pool: oracledb.ConnectionPool, c: dict,
                   intrid: str, validated_desc: str) -> TenureResult:
    res = TenureResult(intrid=intrid, validated_desc=validated_desc)
    binds = {"pid": intrid, "tol": c["tolerance"], "sliver": c["sliver_sqm"]}
    try:
        with pool.acquire() as conn:
            cur = conn.cursor()

            # 1) parcel overlaps
            cur.execute(parcel_overlap_sql(c), binds)
            cols = [d[0].lower() for d in cur.description]
            for row in cur.fetchall():
                rec = dict(zip(cols, row, strict=False))
                if res.tenure_area_ha is None and rec.get("tenure_area_ha") is not None:
                    res.tenure_area_ha = float(rec["tenure_area_ha"])
                ratio = rec.get("coverage_ratio")
                rec["within_y_n"] = "Y" if (ratio is not None and
                                            float(ratio) >= c["within_threshold"]) else "N"
                res.parcels.append(rec)

            # 2) unsurveyed remainder
            cur.execute(remainder_sql(c),
                        {"pid": intrid, "tol": c["tolerance"]})
            r = cur.fetchone()
            res.remainder_ha = float(r[0]) if r and r[0] is not None else 0.0

            # 3) remainder district (only worth querying if there is a remainder)
            if res.remainder_ha and res.remainder_ha >= c["remainder_ha_min"]:
                cur.execute(remainder_district_sql(c),
                            {"pid": intrid, "tol": c["tolerance"]})
                rd = cur.fetchone()
                res.remainder_district = rd[0] if rd else None

            cur.close()
    except Exception as e:           # noqa: BLE001 — log and keep the batch alive
        res.error = f"{type(e).__name__}: {e}"
        log.warning("tenure %s failed: %s", intrid, res.error)
    return res


# =====================================================================
# Output writers
# =====================================================================
def write_xlsx(results: list[TenureResult], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for r in results:
        if r.error or not r.parcels:
            continue
        # Sheet name capped at 31 chars; key by tenure id.
        ws = wb.create_sheet(title=f"T_{r.intrid}"[:31])
        has_remainder = (r.remainder_ha or 0) >= CONFIG["remainder_ha_min"]
        headers = ["OVERLAP_LEGAL_DESCRIPTION", "WITHIN_Y_N", "AREA_HA"]
        if has_remainder:
            headers += ["UNSURVEYED_LAND_DISTRICT", "PHYSICAL_FEATURE"]
        ws.append(headers)
        for i, p in enumerate(r.parcels):
            row = [p.get("legal_desc"), p.get("within_y_n"),
                   r.tenure_area_ha if i == 0 else None]
            if has_remainder:
                # district on first row only; PHYSICAL_FEATURE left blank for a
                # human/GNIS step to fill (reconstruction can't pick the name).
                row += [r.remainder_district if i == 0 else None, None]
            ws.append(row)
    wb.save(path)
    log.info("wrote %s (%d sheets)", path, len(wb.sheetnames))


def write_report(results: list[TenureResult], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["intrid", "n_parcels", "n_within_Y", "n_primary", "n_subdiv",
                    "tenure_area_ha", "remainder_ha", "has_unsurveyed",
                    "remainder_district", "error", "validated_desc"])
        for r in results:
            n_y = sum(1 for p in r.parcels if p.get("within_y_n") == "Y")
            n_prim = sum(1 for p in r.parcels if p.get("parcel_type") == "Primary")
            n_sub = sum(1 for p in r.parcels if p.get("parcel_type") == "Subdivision")
            w.writerow([
                r.intrid, len(r.parcels), n_y, n_prim, n_sub, r.tenure_area_ha,
                r.remainder_ha,
                "Y" if (r.remainder_ha or 0) >= CONFIG["remainder_ha_min"] else "N",
                r.remainder_district, r.error or "",
                (r.validated_desc or "")[:300],
            ])
    log.info("wrote %s", path)


def write_calibration(results: list[TenureResult], path: str) -> None:
    """Per-parcel threshold sweep. For each reconstructed parcel, align it to the
    author's whole/partial treatment (matched by parcel number parsed from the
    validated description), then at each threshold count how many individual
    parcels are classified in agreement. Per-parcel (not per-tenure) so the sweep
    can actually discriminate. Also dumps a per-parcel diagnostic CSV next to it."""
    # Coarse grid across the range, plus a FINE grid near the ceiling: coverage
    # is bimodal with the signal at the top, and the production threshold
    # (0.999) is not expressible on a 0.005 grid — the fine steps let the sweep
    # actually confirm or refute it.
    coarse = [round(0.90 + 0.005 * i, 4) for i in range(21)]      # 0.90..1.00
    fine = [round(0.990 + 0.0005 * i, 4) for i in range(21)]      # 0.990..1.000
    thresholds = sorted(set(coarse + fine))

    # Build the alignable parcel set once: (coverage_ratio, author_treatment)
    aligned = []          # list of (ratio, 'whole'|'partial')
    diag_rows = []        # per-parcel detail for the diagnostic dump
    unmatched = 0         # parcels we couldn't align to the description
    tenures_used = 0
    tenures_skipped = 0   # descriptions the parser flagged UNPARSEABLE
    for r in results:
        if r.error or not r.parcels:
            continue
        treatment = parse_parcel_treatment(r.validated_desc)
        if treatment is UNPARSEABLE or not treatment:
            tenures_skipped += 1
            continue
        tenures_used += 1
        for p in r.parcels:
            ratio = p.get("coverage_ratio")
            if ratio is None:
                continue
            nums = extract_parcel_numbers(p.get("legal_desc"))
            treats = {treatment[n] for n in nums if n in treatment}
            if not treats:
                unmatched += 1
                diag_rows.append((r.intrid, ";".join(sorted(nums)) or "?",
                                  round(float(ratio), 6), "UNMATCHED",
                                  (p.get("legal_desc") or "")[:120]))
                continue
            author = 'partial' if 'partial' in treats else 'whole'
            aligned.append((float(ratio), author))
            diag_rows.append((r.intrid, ";".join(sorted(nums)),
                              round(float(ratio), 6), author,
                              (p.get("legal_desc") or "")[:120]))

    rows = []
    for t in thresholds:
        agree = 0
        for ratio, author in aligned:
            predicted = 'whole' if ratio >= t else 'partial'
            if predicted == author:
                agree += 1
        n = len(aligned)
        rows.append((t, n, agree, round(agree / n, 4) if n else 0.0))

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["threshold", "parcels_aligned", "parcels_agree", "agreement_rate"])
        w.writerows(rows)

    # Diagnostic dump: every parcel's ratio vs the author's treatment, so the
    # whole/partial disagreement can be analyzed (parser noise vs fabric drift).
    diag_path = os.path.splitext(path)[0] + "_perparcel.csv"
    with open(diag_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["intrid", "parcel_numbers", "coverage_ratio",
                    "author_treatment", "parcel_legal_desc"])
        w.writerows(diag_rows)

    log.info("wrote %s", path)
    log.info("wrote %s (per-parcel diagnostic)", diag_path)
    log.info("calibration basis: %d tenures used, %d skipped (unparseable), "
             "%d parcels aligned, %d parcels unmatched",
             tenures_used, tenures_skipped, len(aligned), unmatched)
    if aligned:
        best = max(rows, key=lambda x: x[3])
        log.info("best threshold = %.3f  (per-parcel agreement %.1f%%)",
                 best[0], best[3] * 100)
        whole_ratios = sorted(r for r, a in aligned if a == 'whole')
        part_ratios = sorted(r for r, a in aligned if a == 'partial')
        if whole_ratios:
            below = sum(1 for r in whole_ratios if r < 0.5)
            log.info("  author-WHOLE coverage: min=%.3f median=%.3f (n=%d, %d below 0.5)",
                     whole_ratios[0], whole_ratios[len(whole_ratios)//2],
                     len(whole_ratios), below)
        if part_ratios:
            above = sum(1 for r in part_ratios if r >= 0.999)
            log.info("  author-PARTIAL coverage: max=%.3f median=%.3f (n=%d, %d at/above 0.999)",
                     part_ratios[-1], part_ratios[len(part_ratios)//2],
                     len(part_ratios), above)


# =====================================================================
# Driver
# =====================================================================
def load_tenures(csv_path: str) -> list[tuple[str, str]]:
    out = []
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            out.append((str(row["INTRID_SID"]).strip(),
                        (row.get("LEGAL_DESC") or "").strip()))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Reconstruct overlap sheets + calibrate WITHIN threshold")
    ap.add_argument("--limit", type=int, default=0, help="process only first N tenures (0 = all)")
    ap.add_argument("--threshold", type=float, default=None, help="override WITHIN_Y_N threshold")
    ap.add_argument("--input", type=str, default=None,
                    help="override input CSV path (e.g. the calibration candidates)")
    ap.add_argument("--tag", type=str, default="",
                    help="suffix for output filenames, e.g. --tag calib")
    args = ap.parse_args()
    if args.threshold is not None:
        CONFIG["within_threshold"] = args.threshold
    if args.input:
        CONFIG["input_csv"] = args.input
    if args.tag:
        for k in ("out_xlsx", "out_report", "out_calib"):
            base, ext = os.path.splitext(CONFIG[k])
            CONFIG[k] = f"{base}_{args.tag}{ext}"

    if not CONFIG["password"]:
        log.error("Set BCGW_PASSWORD (and BCGW_USER/BCGW_DSN) in the environment.")
        return 2

    tenures = load_tenures(CONFIG["input_csv"])
    if args.limit:
        tenures = tenures[:args.limit]
    log.info("loaded %d tenures", len(tenures))

    pool = oracledb.create_pool(
        user=CONFIG["user"], password=CONFIG["password"], dsn=CONFIG["dsn"],
        min=2, max=CONFIG["max_workers"], increment=1,
    )
    log.info("connection pool up (max=%d)", CONFIG["max_workers"])

    results: list[TenureResult] = []
    try:
        with ThreadPoolExecutor(max_workers=CONFIG["max_workers"]) as ex:
            futs = {ex.submit(process_tenure, pool, CONFIG, tid, desc): tid
                    for tid, desc in tenures}
            done = 0
            for fut in as_completed(futs):
                results.append(fut.result())
                done += 1
                if done % 20 == 0:
                    log.info("  %d/%d tenures done", done, len(tenures))
    finally:
        pool.close()

    results.sort(key=lambda r: r.intrid)
    write_xlsx(results, CONFIG["out_xlsx"])
    write_report(results, CONFIG["out_report"])
    write_calibration(results, CONFIG["out_calib"])

    n_ok = sum(1 for r in results if not r.error and r.parcels)
    n_err = sum(1 for r in results if r.error)
    n_empty = sum(1 for r in results if not r.error and not r.parcels)
    log.info("done. %d reconstructed, %d empty (no parcel overlap), %d errored",
             n_ok, n_empty, n_err)
    return 0


if __name__ == "__main__":
    sys.exit(main())
